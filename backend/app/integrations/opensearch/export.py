from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.timezone import to_baku


_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header(ws: Any, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _auto_width(ws: Any, max_width: int = 45) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_rows(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    _style_header(ws, len(headers))
    for row in rows:
        ws.append(row)
    _auto_width(ws)


def build_excel_export(
    ip: str,
    logs_data: dict[str, Any],
    summary_data: dict[str, Any] | None = None,
    full_data: dict[str, Any] | None = None,
) -> io.BytesIO:
    wb = Workbook()

    # --- Sheet 1: Logs ---
    ws_logs = wb.active
    ws_logs.title = "Logs"
    log_headers = [
        "Timestamp", "Source", "Index", "Doc ID",
        "Src IP", "Src Port", "Dst IP", "Dst Port",
        "Protocol", "Action", "Application", "Rule", "Policy",
        "User", "Domain", "URL", "Bytes", "Packets", "Direction",
    ]
    log_rows = []
    for log in logs_data.get("logs", []):
        log_rows.append([
            to_baku(log.get("timestamp", "")), log.get("source_name", ""), log.get("index", ""), log.get("id", ""),
            log.get("source_ip", ""), log.get("source_port", ""), log.get("destination_ip", ""), log.get("destination_port", ""),
            log.get("protocol", ""), log.get("action", ""), log.get("application", ""), log.get("rule", ""), log.get("policy", ""),
            log.get("user", ""), log.get("domain", ""), log.get("url", ""),
            log.get("bytes", ""), log.get("packets", ""), log.get("direction", ""),
        ])
    _write_rows(ws_logs, log_headers, log_rows)

    # --- Sheet 2: Summary ---
    if summary_data and summary_data.get("status", {}).get("status") == "ok":
        ws_summary = wb.create_sheet("Summary")
        summary_rows = [
            ["IP", ip],
            ["Window", summary_data.get("window", "")],
            ["Internal Connections", summary_data.get("internal_connections", 0)],
            ["External Connections", summary_data.get("external_connections", 0)],
            ["Security Events", summary_data.get("security_events", 0)],
            ["Users", ", ".join(summary_data.get("users", []))],
        ]
        for src, count in summary_data.get("source_stats", {}).items():
            summary_rows.append([f"Source: {src}", count])
        for idx, count in summary_data.get("index_stats", {}).items():
            summary_rows.append([f"Index: {idx}", count])
        _write_rows(ws_summary, ["Metric", "Value"], summary_rows)

    # --- Full aggregation data ---
    if full_data:
        # Sheet: Domains
        ws_dom = wb.create_sheet("Domains")
        dom_rows = []
        for b in full_data.get("domains", {}).get("buckets", []):
            key = b.get("key", {})
            first_seen = to_baku(b.get("first_seen", {}).get("value_as_string", ""))[:19]
            last_seen = to_baku(b.get("last_seen", {}).get("value_as_string", ""))[:19]
            dom_rows.append([
                key.get("domain", ""),
                key.get("application", ""),
                b.get("doc_count", 0),
                first_seen,
                last_seen,
            ])
        _write_rows(ws_dom, ["Domain", "Application", "Count", "First Seen", "Last Seen"], dom_rows)

        # Sheet: Top IPs
        ws_ips = wb.create_sheet("Top IPs")
        ips_data = full_data.get("ips", {})
        ip_rows = []
        for item in ips_data.get("as_source", [])[:50]:
            ip_rows.append([item["key"], "source", item["doc_count"]])
        for item in ips_data.get("as_destination", [])[:50]:
            ip_rows.append([item["key"], "destination", item["doc_count"]])
        _write_rows(ws_ips, ["IP", "Role", "Count"], ip_rows)

        # Sheet: Ports
        ws_ports = wb.create_sheet("Ports")
        port_rows = [[p["key"], p["doc_count"]] for p in full_data.get("ports", [])[:50]]
        _write_rows(ws_ports, ["Port", "Count"], port_rows)

        # Sheet: Protocols
        ws_proto = wb.create_sheet("Protocols")
        proto_rows = [[p["key"], p["doc_count"]] for p in full_data.get("protocols", [])]
        _write_rows(ws_proto, ["Protocol", "Count"], proto_rows)

        # Sheet: Actions
        ws_act = wb.create_sheet("Actions")
        act_rows = [[a["key"], a["doc_count"]] for a in full_data.get("actions", [])]
        _write_rows(ws_act, ["Action", "Count"], act_rows)

        # Sheet: Users
        ws_users = wb.create_sheet("Users")
        user_rows = [[u["key"], u["doc_count"]] for u in full_data.get("users", [])[:50]]
        _write_rows(ws_users, ["User", "Count"], user_rows)

    ws_logs.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
