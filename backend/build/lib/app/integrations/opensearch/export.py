from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def _style_header(ws: Any, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _auto_width(ws: Any, max_width: int = 45) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _write_rows(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    _style_header(ws, len(headers))
    for row in rows:
        ws.append(row)
    _auto_width(ws)


def build_excel_export(
    ip: str,
    logs_data: dict[str, Any],
    summary_data: dict[str, Any] | None = None,
    domains_data: dict[str, Any] | None = None,
) -> io.BytesIO:
    wb = Workbook()

    # --- Sheet 1: Logs ---
    ws_logs = wb.active
    ws_logs.title = "Logs"

    log_headers = [
        "Timestamp", "Source", "Index", "Doc ID",
        "Src IP", "Src Port", "Dst IP", "Dst Port",
        "Protocol", "Action", "Application", "Rule", "Policy",
        "User", "Domain", "URL", "Bytes", "Packets", "Direction",
    ]

    log_rows = []
    for log in logs_data.get("logs", []):
        log_rows.append([
            log.get("timestamp", ""),
            log.get("source_name", ""),
            log.get("index", ""),
            log.get("id", ""),
            log.get("source_ip", ""),
            log.get("source_port", ""),
            log.get("destination_ip", ""),
            log.get("destination_port", ""),
            log.get("protocol", ""),
            log.get("action", ""),
            log.get("application", ""),
            log.get("rule", ""),
            log.get("policy", ""),
            log.get("user", ""),
            log.get("domain", ""),
            log.get("url", ""),
            log.get("bytes", ""),
            log.get("packets", ""),
            log.get("direction", ""),
        ])

    _write_rows(ws_logs, log_headers, log_rows)

    # --- Sheet 2: Summary ---
    if summary_data and summary_data.get("status", {}).get("status") == "ok":
        ws_summary = wb.create_sheet("Summary")
        summary_headers = ["Metric", "Value"]
        summary_rows = [
            ["IP", ip],
            ["Window", summary_data.get("window", "")],
            ["Internal Connections", summary_data.get("internal_connections", 0)],
            ["External Connections", summary_data.get("external_connections", 0)],
            ["Security Events", summary_data.get("security_events", 0)],
            ["Users", ", ".join(summary_data.get("users", []))],
            ["Top Internal Destinations", len(summary_data.get("top_internal_destinations", []))],
            ["Top External Destinations", len(summary_data.get("top_external_destinations", []))],
            ["Top Internal Ports", len(summary_data.get("top_internal_ports", []))],
            ["Top External Ports", len(summary_data.get("top_external_ports", []))],
            ["Top Domains", len(summary_data.get("top_domains", []))],
        ]

        for src, count in summary_data.get("source_stats", {}).items():
            summary_rows.append([f"Source: {src}", count])
        for idx, count in summary_data.get("index_stats", {}).items():
            summary_rows.append([f"Index: {idx}", count])

        _write_rows(ws_summary, summary_headers, summary_rows)

        # --- Sheet 3: Internal Destinations ---
        ws_int = wb.create_sheet("Internal Destinations")
        _write_rows(ws_int, ["IP", "Port", "Count"], [
            [d.get("ip", ""), d.get("port", ""), d.get("count", 0)]
            for d in summary_data.get("top_internal_destinations", [])
        ])

        # --- Sheet 4: External Destinations ---
        ws_ext = wb.create_sheet("External Destinations")
        _write_rows(ws_ext, ["IP", "Port", "Count"], [
            [d.get("ip", ""), d.get("port", ""), d.get("count", 0)]
            for d in summary_data.get("top_external_destinations", [])
        ])

        # --- Sheet 5: Ports ---
        ws_ports = wb.create_sheet("Ports")
        port_rows = []
        for p in summary_data.get("top_internal_ports", []):
            port_rows.append([p.get("port", ""), "internal", p.get("count", 0)])
        for p in summary_data.get("top_external_ports", []):
            port_rows.append([p.get("port", ""), "external", p.get("count", 0)])
        _write_rows(ws_ports, ["Port", "Direction", "Count"], port_rows)

        # --- Sheet 6: Domains ---
        ws_domains = wb.create_sheet("Domains")
        _write_rows(ws_domains, ["Domain", "Count"], [
            [d.get("ip", ""), d.get("count", 0)]
            for d in summary_data.get("top_domains", [])
        ])

    # --- Sheet 7: Domain Activity (from aggregation) ---
    if domains_data and domains_data.get("buckets"):
        ws_dom_activity = wb.create_sheet("Domain Activity")
        dom_activity_rows = []
        for bucket in domains_data["buckets"]:
            domain = bucket.get("key", {}).get("domain", "")
            application = bucket.get("key", {}).get("application", "")
            count = bucket.get("doc_count", 0)
            first_seen = bucket.get("first_seen", {}).get("value_as_string", "")
            last_seen = bucket.get("last_seen", {}).get("value_as_string", "")
            dom_activity_rows.append([domain, application, count, first_seen, last_seen])
        _write_rows(ws_dom_activity, ["Domain", "Application", "Count", "First Seen", "Last Seen"], dom_activity_rows)

    # Freeze panes on logs sheet
    ws_logs.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
